from datetime import datetime, timedelta

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# loading sheet
src = openpyxl.load_workbook("table.xlsx")
src_sheet = src["sheet_name"]


START_DATE_COL = 0
END_DATE_COL = 1
ENERGY_COL = 2


def end_date(sheet: Worksheet, start_row: int) -> [str, int]:
    """Finds last row where energy value is 0 and returns [Date string, row index]"""

    for row_idx in range(start_row, sheet.max_row):

        # if energy is not zero, we passed the last value, subtract 1
        # NOTE: cell index start at 1, so need to add 1.
        if sheet.cell(row_idx + 1, ENERGY_COL + 1).value != 0:
            return [sheet.cell(row_idx, END_DATE_COL + 1).value, row_idx]
        else:
            end_date(sheet, start_row + 1)


sheet_iter = enumerate(src_sheet.iter_rows())

# skip column names
next(sheet_iter)

for idx, row in sheet_iter:

    if row[ENERGY_COL] != 0:
        start = row[START_DATE_COL].value
        end, end_idx = end_date(src_sheet, idx + 1)

        duration = datetime.fromisoformat(end) - datetime.fromisoformat(start)
        print(f"Start: {start}, End: {end}, Duration: {duration}")

        # consume iterator
        for _ in range(end_idx - idx):
            next(sheet_iter)
